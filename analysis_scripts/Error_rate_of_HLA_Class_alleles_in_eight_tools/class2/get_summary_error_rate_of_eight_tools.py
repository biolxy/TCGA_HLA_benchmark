#!/usr/bin/env python
# -*- coding: utf-8 -*-
u"""
File Name   : get_summary_error_rate_of_eight_tools.py .

Author      : biolxy
E-mail      : biolxy@aliyun.com
Created Time: 2020-07-20 23:30:12
version     : 1.0
Function    : The author is too lazy to write nothing
Usage       :
"""
import sys
import os
import pandas as pd
from openpyxl import load_workbook


def get_df_name(infile):
    df = pd.DataFrame(pd.read_csv(infile, encoding='utf-8', sep='\t'))
    name = infile.split(".")[1]
    return df, name


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    from openpyxl import load_workbook
    import pandas as pd
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        # if startrow is None and sheet_name in writer.book.sheetnames:
        #     startrow = writer.book[sheet_name].max_row
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()



def main():
    """
    https://blog.csdn.net/qq_38486203/article/details/90263558
    """
    xlsxfile = 'summary_class2_error_rate_of_eight_tools.xlsx'

    for error_rate_file in sys.argv[1:]:
        df, name = get_df_name(error_rate_file)
        append_df_to_excel(xlsxfile, df, sheet_name=name, index=False)







if __name__ == "__main__":
    main()